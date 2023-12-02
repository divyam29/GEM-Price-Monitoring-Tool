import math
import multiprocessing
import os
import time
from bs4 import BeautifulSoup
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl.styles import Font
from timeit import default_timer as timer


def slice_list(lst):
    k, m = divmod(len(lst), 4)
    return (lst[i * k + min(i, m) : (i + 1) * k + min(i + 1, m)] for i in range(4))


def s_webdriver(lst, return_dict, lst_name):
    URL = "https://mkp.gem.gov.in/41111700-microscopes-pathological-and-research/search#/?sort_type=price_in_asc&_xhr=1"

    options = webdriver.ChromeOptions()
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--incognito")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu") if os.name == "nt" else None
    driver = webdriver.Chrome(options=options)
    driver.get(URL)

    ct=1
    names = []
    for i1 in lst:
        driver.get(i1)
        time.sleep(4)
        html = driver.page_source

        print(f"{ct}: ",end="")
        name = {}
        name["link"] = i1
        name["name"] = "-"
        name["curr_price"] = "-"
        name["ESAW-Title"] = ""
        name["ESAW-price"] = ""
        name["ESAW-productid"] = ""
        try:
            soup = BeautifulSoup(html, "lxml")
            parentul = soup.find("ul", attrs={"id": "search-result-items"})
            parentli = parentul.find("li", attrs={"class": "clearfix"})
            titlediv = parentli.find("div", attrs={"class": "variant-desc"})
            titlespan = titlediv.find("span", attrs={"class": "variant-title"})
            title = titlespan.a.text
            print(title)

            parent_price = titlediv.find("span", attrs={"variant-final-price"})
            curr_price = parent_price.span.text
            print(curr_price)

            name["name"] = title
            name["curr_price"] = curr_price
            name["ESAW-Title"] = ""
            name["ESAW-price"] = ""
            name["ESAW-productid"] = ""

            if str(title).find("ESAW") == -1 and str(title).find("E.S.A.W") == -1:
                try:
                    x = i1
                    n = x.find("search#/?")
                    url2 = x[: n + 9] + "q=ESAW&" + x[n + 9 :]
                    driver.get(url2)
                    time.sleep(4)
                    parent_ele = driver.find_element(By.CLASS_NAME, "variant-title")
                    ele_list = parent_ele.find_elements(By.XPATH, "*")
                    wait = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((ele_list[0]))
                    )
                    ele_list[0].click()
                    time.sleep(5)
                    h_title = driver.find_element(By.CLASS_NAME, "like-h3")
                    h_title_text = h_title.text
                    print(h_title_text)
                    h_price = driver.find_element(By.CLASS_NAME, "m-w")
                    h_price_text = h_price.text
                    print(h_price_text)
                    h_product_id = driver.find_element(By.CLASS_NAME, "item_sku")
                    h_product_id_text = h_product_id.text
                    print(h_product_id_text)
                    name["ESAW-Title"] = h_title_text
                    name["ESAW-price"] = h_price_text
                    name["ESAW-productid"] = h_product_id_text

                except:
                    name["ESAW-Title"] = "N/A"
                    name["ESAW-price"] = "N/A"
                    name["ESAW-productid"] = "N/A"
                    print()

        except:
            print()

        names.append(name)
        print()
        ct+=1

    driver.quit()
    return_dict[lst_name] = names


def convert_to_excel(names1, filename):
    names = []
    for d1 in names1:
        for d2 in d1:
            names.append(d2)

    filename = f"ZZZOUTPUT/{filename}_output.xlsx"
    df = pd.DataFrame.from_dict(names)
    df.to_excel(filename)
    esaw = []
    not_esaw = []

    wrkbk = openpyxl.load_workbook(filename)
    wb = wrkbk.active

    for i in range(2, wb.max_row + 1):
        cell_obj = wb.cell(row=i, column=3)
        cell_name = str(cell_obj.value)
        if cell_name.find("ESAW") == -1 and cell_name.find("E.S.A.W") == -1:
            not_esaw.append(int(cell_obj.coordinate[1:]))
        else:
            esaw.append(int(cell_obj.coordinate[1:]))

    for i in esaw:
        for rows in wb.iter_rows(min_row=i, max_row=i, min_col=3):
            for cell in rows:
                cell.font = Font(color="003366FF")

    for i in not_esaw:
        for rows in wb.iter_rows(min_row=i, max_row=i, min_col=3):
            for cell in rows:
                cell.font = Font(color="00FF0000")

    wb.column_dimensions["C"].width = 40
    wb.column_dimensions["D"].width = 12
    wb.column_dimensions["E"].width = 40
    wb.column_dimensions["F"].width = 12
    wb.column_dimensions["G"].width = 20

    wrkbk.save(filename)

def main(filename):
    start = timer()

    path = "ZZZOUTPUT"
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)
        print(f"The new directory '{path}' is created!\n")

    try:
        f = open(filename, "r")
    except OSError:
        print(f"Unable to open file")

    data = pd.read_excel(filename)
    
    print(f"Excel File: '{filename}'\n")

    x = filename.rfind("/")
    filename = filename[x + 1 : -5]

    urllist = data["LINKS"].tolist()
    print(f"Number of Links to Process: {len(urllist)}")
    lst1, lst2, lst3, lst4 = slice_list(urllist)
    print(f"Number of Links in Thread 1: {len(lst1)}")
    print(f"Number of Links in Thread 2: {len(lst2)}")
    print(f"Number of Links in Thread 3: {len(lst3)}")
    print(f"Number of Links in Thread 4: {len(lst4)}")

    manager = multiprocessing.Manager()
    return_dict = manager.dict()
    jobs = []

    p = multiprocessing.Process(target=s_webdriver, args=(lst1, return_dict, "lst1"))
    jobs.append(p)
    p.start()
    p = multiprocessing.Process(target=s_webdriver, args=(lst2, return_dict, "lst2"))
    jobs.append(p)
    p.start()
    p = multiprocessing.Process(target=s_webdriver, args=(lst3, return_dict, "lst3"))
    jobs.append(p)
    p.start()
    p = multiprocessing.Process(target=s_webdriver, args=(lst4, return_dict, "lst4"))
    jobs.append(p)
    p.start()

    for proc in jobs:
        proc.join()

    names1 = return_dict.values()
    convert_to_excel(names1, filename)
    end = timer()

    secs = end-start
    print(f"Time Taken: \n{int(secs//60)} Minutes and {math.floor(secs%60)} Seconds\n")

    filepath = f"OUTPUT/{filename}_output.xlsx"
    print(f"Processed Excel File: '{os.path.abspath(filepath)}'\n")


# pyinstaller --noconfirm --onedir --console  .\main.py
